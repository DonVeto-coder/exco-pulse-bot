"""
OEB Exco Action Tracker — data engine.
Reads the workbook and computes every figure the dashboard needs.
Mirrors the logic in OEB_DashboardFormulas (SUMPRODUCT on tbl_ExcoActions).
"""
import openpyxl, warnings, datetime, json, sys
from openpyxl.utils import range_boundaries
from collections import Counter, defaultdict
warnings.filterwarnings('ignore')

# role title -> dashboard label (as shown in the email)
ROLE_LABEL = {
    "Chief Executive Officer":   ("CEO",        "Warwick"),
    "Finance Manager":           ("Finance",    "Monique"),
    "Human Resources Manager":   ("HR",         "Natasha"),
    "Sales Manager":             ("Sales",      "Donovan"),
    "Production Manager":        ("Production", "Rob"),
    "Innovation Manager":        ("Innovation", "Rob + Warwick"),
    "Marketing Manager":         ("Marketing",  "Nic"),
    "IT Manager":                ("IT",         "Nic"),
    "Strategy Manager":          ("Strategy",   "Nic"),
}

def as_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    return None

def norm_status(v):
    s = str(v or "").strip().lower()
    if s.startswith("done"): return "Done"
    if s in ("merged",): return "Merged"
    if "progress" in s or "going" in s: return "In Progress"
    if s in ("", "please select", "none"): return "Unset"
    return "Not Started"

def build(xlsm_path, report_date=None):
    if report_date is None:
        report_date = datetime.date.today()
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws = wb["Actions"]
    ref = ws.tables["tbl_ExcoActions"]
    ref = ref.ref if hasattr(ref, "ref") else ref
    minc, minr, maxc, maxr = range_boundaries(ref)
    hdr = [ws.cell(minr, c).value for c in range(minc, maxc+1)]
    def ci(name):
        nl = name.lower()
        # prefer an exact (stripped, case-insensitive) header match
        for i, h in enumerate(hdr):
            if h and str(h).replace("\n", " ").strip().lower() == nl: return minc + i
        # fall back to substring match
        for i, h in enumerate(hdr):
            if h and nl in str(h).replace("\n", " ").lower(): return minc + i
        return None
    C_status = ci("Status"); C_owner = ci("Owner"); C_eff = ci("Effective Due"); C_closed = ci("Date Closed")

    total_open = overdue = todo = inprog = 0
    owner_over = Counter(); owner_open = Counter(); owner_total = Counter()
    for r in range(minr+1, maxr+1):
        owner = str(ws.cell(r, C_owner).value or "").strip()
        if owner in ROLE_LABEL:
            owner_total[owner] += 1     # all actions for this owner (any status)
        st = norm_status(ws.cell(r, C_status).value)
        if st in ("Done", "Merged", "Unset"):
            continue
        total_open += 1
        owner_open[owner] += 1
        ed = as_date(ws.cell(r, C_eff).value)
        raw_ed = str(ws.cell(r, C_eff).value or "")
        is_overdue = (ed is not None and ed < report_date) or ("missing" in raw_ed.lower())
        if is_overdue:
            overdue += 1; owner_over[owner] += 1
        elif st == "In Progress":
            inprog += 1
        else:
            todo += 1

    owners = []
    for role, cnt in owner_over.most_common():
        if role in ROLE_LABEL:
            dept, who = ROLE_LABEL[role]
            owners.append({"role": role, "dept": dept, "who": who,
                           "overdue": cnt, "open": owner_open.get(role, 0)})

    # ---- monthly delivery pulse per owner (last 5 calendar months ending report month) ----
    # planned[m] = actions with Effective Due Date in month m
    # delivered[m] = Done actions with Date Closed in month m   (verified vs reference)
    mlabels = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
    months = []
    y, mo = report_date.year, report_date.month
    for back in range(4, -1, -1):
        yy = y; mm = mo - back
        while mm <= 0: mm += 12; yy -= 1
        months.append((yy, mm))
    planned = {r: {ym: 0 for ym in months} for r in ROLE_LABEL}
    delivered = {r: {ym: 0 for ym in months} for r in ROLE_LABEL}
    for r in range(minr+1, maxr+1):
        role = str(ws.cell(r, C_owner).value or "").strip()
        if role not in ROLE_LABEL: continue
        ed = as_date(ws.cell(r, C_eff).value)
        dc = as_date(ws.cell(r, C_closed).value)
        st = norm_status(ws.cell(r, C_status).value)
        if ed and (ed.year, ed.month) in planned[role]:
            planned[role][(ed.year, ed.month)] += 1
        if st == "Done" and dc and (dc.year, dc.month) in delivered[role]:
            delivered[role][(dc.year, dc.month)] += 1
    pulses = {}
    cur_ym = (report_date.year, report_date.month)
    for role in ROLE_LABEL:
        pulses[role] = [
            {"label": mlabels[mm-1], "delivered": delivered[role][(yy, mm)],
             "planned": planned[role][(yy, mm)], "current": (yy, mm) == cur_ym}
            for (yy, mm) in months]

    return {
        "report_date": report_date.isoformat(),
        "waffle": {"open": total_open, "todo": todo, "inprog": inprog, "overdue": overdue},
        "owners_overdue": owners,
        "owner_open": dict(owner_open),
        "owner_total": dict(owner_total),
        "pulses": pulses,
    }

if __name__ == "__main__":
    rd = datetime.date(2026, 6, 15) if len(sys.argv) < 2 else datetime.date.fromisoformat(sys.argv[1])
    data = build("/root/exco/tracker.xlsm", rd)
    print(json.dumps(data, indent=2))
